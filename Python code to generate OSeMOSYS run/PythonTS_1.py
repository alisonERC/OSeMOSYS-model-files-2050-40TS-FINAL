#initial stuff that doesn't change
#~ alison$ cd /Users/alison/Documents/2021/2021\ OSeMOSYS\ model 
#glpsol -m OSeMOSYS_2017_11_08.txt -d DRC_Input_data_West_v3.txt -o West_v3_1.txt 
#NB   set file path to read from and file path to write to

file2 = open(r"/Users/alison/Documents/2022/2021 OSeMOSYS model/WandS_Sc8s.txt","w+")


print("#*************",file=file2)

print("param	AnnualExogenousEmission	         default	0	:=	;\n",
"param	AnnualEmissionLimit	         default	99999	:=	;\n",
"param	ModelPeriodExogenousEmission	 default	0	:=	;\n",
"param	ModelPeriodEmissionLimit	 default	99999	:=	;\n",
#    "param      TotalAnnualMaxCapacityInvestment default	99999	:=	;\n",- defined later
#    "param	TotalAnnualMinCapacityInvestment default	0	:=	;\n",- defined later
"param	TotalTechnologyAnnualActivityUpperLimit	default	99999	:=	;\n",
"param	TotalTechnologyAnnualActivityLowerLimit	default	0	:=	;\n",
"param	TotalTechnologyModelPeriodActivityUpperLimit	default	99999	:=	;\n",
"param	TotalTechnologyModelPeriodActivityLowerLimit	default	0	:=	;\n",
"param	RETagTechnology	                 default	0	:=	;\n",
"param	RETagFuel	                 default	0	:=	;\n",
"param	REMinProductionTarget	         default	0	:=	;\n",
"param	EmissionActivityRatio	         default	0	:=	;\n",
"param	AccumulatedAnnualDemand	         default	0	:=	;",	
file = file2)



#other non variable parameters
print("param 	DiscountRate	                 default	0.1	:=	;", file = file2)
print("param 	DepreciationMethod	         default	1	:=	;", file = file2)

print("#*************",file=file2)

#define storage sets
print("set      SEASON                           :=         1            ;", file = file2)
print("set      DAYTYPE                          :=         1            ;", file = file2)
print("set      DAILYTIMEBRACKET	         :=         1	         ;", file = file2)
print("set      STORAGE			         :=		         ;", file = file2)

# user defined sets not in spreadsheet (EMISSION, MODE OF OPERATION,REGION) these should remain the same  

print("set      EMISSION		         :=	    CO2	NOX	        ;", file = file2)
print("set      MODE_OF_OPERATION	         :=	    1 	                ;",file=file2)
print("set      REGION	                         :=	    DRC 	        ;",file=file2)

print("#*************",file=file2)

# user defined sets in spreadsheet (TECHNOLOGY,FUEL, YEAR, TIMESLICE)  data read from excel

import openpyxl
path="/Users/alison/Documents/2022/2021 OSeMOSYS model/DRC_input_data_West_v350_4_10_s.xlsx"
path2="/Users/alison/Documents/2022/2021 OSeMOSYS model/DRC_input_data_South_v350_4_10_s.xlsx"
# To open the workbook
wb_obj = openpyxl.load_workbook(path, data_only=True)
wb_obj2 = openpyxl.load_workbook(path2, data_only=True)
#define years
Yearsheet_obj=wb_obj['YEAR']
Baseyear=Yearsheet_obj['C2'].value
Yearnumber=Yearsheet_obj.cell(row=3,column=3).value
print(Yearnumber)
Years=list(map(str, range(Baseyear,Baseyear+Yearnumber )))
print("set  YEAR  :=  ",*Years, sep = "    ",end='',file=file2),
print("     ;",file=file2)
#define Timeslices
TSsheet_obj=wb_obj['TS_DRC']
max_row_TSsheet = max((c.row for c in TSsheet_obj['O'] if c.value is not None))
print(max_row_TSsheet)
result = ''
result += 'set ' + ' TIMESLICE ' + ' := '
print(result)
for row in TSsheet_obj.iter_rows(min_row=5, max_row=max_row_TSsheet,min_col=16,max_col=16):
    result += " ".join([data.value for data in row]) + " "
#    print(result)
result += ";\n"
print(result,file=file2)

#Define technology
Techsheet_obj=wb_obj['TECHNOLOGY']
Techsheet_obj2=wb_obj2['TECHNOLOGY']
max_row_Techsheet = max((c.row for c in Techsheet_obj['D'] if c.value is not None))
max_row_Techsheet2 = max((c.row for c in Techsheet_obj2['D'] if c.value is not None))
result = ''
result += 'set ' + ' TECHNOLOGY ' + ' := '
for row in Techsheet_obj.iter_rows(min_row=2,max_row=max_row_Techsheet ,min_col=4,max_col=4):
    result += " ".join([data.value for data in row]) + " "    
for row in Techsheet_obj2.iter_rows(min_row=2,max_row=max_row_Techsheet2 ,min_col=4,max_col=4):
    result += " ".join([data.value for data in row]) + " "    
result += ";\n"
print(result,file=file2)

#Define fuels
Fuelsheet_obj=wb_obj['FUEL']
Fuelsheet_obj2=wb_obj2['FUEL']

max_row_Fuelsheet = max((c.row for c in Fuelsheet_obj['D'] if c.value is not None))
max_row_Fuelsheet2 = max((c.row for c in Fuelsheet_obj2['D'] if c.value is not None))

result = ''
result += 'set ' + ' FUEL ' + ' := '
for row in Fuelsheet_obj.iter_rows(min_row=2,max_row=max_row_Fuelsheet ,min_col=4,max_col=4):
    result += " ".join([data.value for data in row]) + " "    
for row in Fuelsheet_obj2.iter_rows(min_row=2,max_row=max_row_Fuelsheet2 ,min_col=4,max_col=4):
    result += " ".join([data.value for data in row]) + " "    
result += ";\n"
print(result,file=file2)


print("#*************",file=file2)


#Define sets for printing
print("set FUELd := D1RES D1IND D1OTH D2RES D2IND D2OTH;", file = file2)
print("set FUELe := D1ELC01 D2ELC01           ;",file=file2)
print("set TRAN  :=D1PWRTRN D2PWRTRN    ;",file=file2)

print("#*************",file=file2)



#Read TS information
TSsheet_obj=wb_obj['TS_DRC']
max_row_TS = max((c.row for c in TSsheet_obj['O'] if c.value is not None))
print("param       YearSplit      :   ",*Years, sep = "    ",end='',file=file2)
print("    :=",file=file2) 

for i in range(5, max_row_TS + 1):
    TSname = TSsheet_obj.cell(row = i, column = 16)
    TSshare = TSsheet_obj.cell(row = i, column = 17)
    print(TSname.value+"    ",(str(TSshare.value)+"    ")*Yearnumber,file=file2),
print("    ;",file=file2)

print("#*************",file=file2)

# Read Demand information
Demandsheet_obj=wb_obj['Demand']
Demandsheet_obj2=wb_obj2['Demand']

max_row_Demand = max((c.row for c in Demandsheet_obj['D'] if c.value is not None))
max_row_Demand2 = max((c.row for c in Demandsheet_obj2['D'] if c.value is not None))

print("param     SpecifiedAnnualDemand	         default	0	:=     ",file=file2)
print("[DRC,*,*]         :   ",*Years, sep = "    ",end='',file=file2)
print("    :=",file=file2) 

for row in  Demandsheet_obj.iter_rows(min_row=4,max_row=max_row_Demand ,min_col=4,max_col=4+Yearnumber):
    result = ''


    for cell in row:
        result += " "+ str(cell.value) + " "

    print(result,file=file2)
for row in  Demandsheet_obj2.iter_rows(min_row=4,max_row=max_row_Demand2 ,min_col=4,max_col=4+Yearnumber):
    result = ''


    for cell in row:
        result += " "+ str(cell.value) + " "

    print(result,file=file2)
#print("\n",file=file2)
print("    ;",file=file2)
print("#*************",file=file2)
### 
print('end demand information')

###
print("param     SpecifiedDemandProfile	         default	0	:=    ",file=file2)
Demandprofilesheet_obj=wb_obj['Demand profile']
Demandprofilesheet_obj2=wb_obj2['Demand profile']

max_row_Demandprofile = max((c.row for c in Demandprofilesheet_obj['E'] if c.value is not None))
max_row_Demandprofile2 = max((c.row for c in Demandprofilesheet_obj2['E'] if c.value is not None))

for row in  Demandprofilesheet_obj.iter_rows(min_row=4,max_row=max_row_Demandprofile ,min_col=5,max_col=5+Yearnumber):
    result = ''
    for cell in row:

        result += " "+ str(cell.value) + " "
    print(result,file=file2)    
for row in  Demandprofilesheet_obj2.iter_rows(min_row=4,max_row=max_row_Demandprofile2 ,min_col=5,max_col=5+Yearnumber):
    result = ''
    for cell in row:

        result += " "+ str(cell.value) + " "

    print(result,file=file2)
print("    ;",file=file2)
print("#*************",file=file2)
print(cell.row)



# Read Input to technologies (defines efficiency)
print("param     InputActivityRatio	          default	0	:=    ",file=file2)
Inputsheet_obj=wb_obj['InputActivityRatio']
Inputsheet_obj2=wb_obj2['InputActivityRatio']

max_row_Input = max((c.row for c in Inputsheet_obj['A'] if c.value is not None))
max_row_Input2 = max((c.row for c in Inputsheet_obj2['A'] if c.value is not None))

for i in range(2, max_row_Input+1):
    result = ""
    print(Inputsheet_obj.cell(row = i, column = 4).value,*Years, sep = "  ",end="",file=file2)
    print("    := ",file=file2) 
    result +=Inputsheet_obj.cell(row = i, column = 5).value 
    for j in range (6,6+Yearnumber):
        result +=  "   "+ str(Inputsheet_obj.cell(row=i,column=j).value) + "   "    
    print(result,file=file2)

for i in range(2, max_row_Input2+1):
    result = ""
    print(Inputsheet_obj2.cell(row = i, column = 4).value,*Years, sep = "  ",end="",file=file2)
    print("    := ",file=file2) 
    result +=Inputsheet_obj2.cell(row = i, column = 5).value 
    for j in range (6,6+Yearnumber):
        result +=  "   "+ str(Inputsheet_obj2.cell(row=i,column=j).value) + "   "    
    print(result,file=file2)
    
print("    ;",file=file2)
print("#*************",file=file2)
print(cell.row)

#Read Output from technologies (defines efficiency)
print("param     OutputActivityRatio	        default	0	:=    ",file=file2)
Outputsheet_obj=wb_obj['OutputActivityRatio']
Outputsheet_obj2=wb_obj2['OutputActivityRatio']

max_row_Output = max((c.row for c in Outputsheet_obj['A'] if c.value is not None))
max_row_Output2 = max((c.row for c in Outputsheet_obj2['A'] if c.value is not None))

for i in range(2, max_row_Output+1):
    result = ""
    print(Outputsheet_obj.cell(row = i, column = 4).value,*Years, sep = "  ",end="",file=file2)
    print("    :=",file=file2) 
    result +=Outputsheet_obj.cell(row = i, column = 5).value 
    for j in range (6,6+Yearnumber):
        result +=  "   "+ str(Outputsheet_obj.cell(row=i,column=j).value) + "   "    
    print(result,file=file2)
for i in range(2, max_row_Output2+1):
    result = ""
    print(Outputsheet_obj2.cell(row = i, column = 4).value,*Years, sep = "  ",end="",file=file2)
    print("    :=",file=file2) 
    result +=Outputsheet_obj2.cell(row = i, column = 5).value 
    for j in range (6,6+Yearnumber):
        result +=  "   "+ str(Outputsheet_obj2.cell(row=i,column=j).value) + "   "    
    print(result,file=file2)
print("    ;",file=file2)
print("#*************",file=file2)


#costs (Fixed cost, Captial cost,  Residual capacity) all one variable
#Put loop here to do this over the three spreadsheets becasue they are the same

sheetNames = wb_obj.sheetnames
for i in range(len(sheetNames)):
    if (sheetNames[i] in ['FixedCost', 'CapitalCost','ResidualCapacity']):
        print('param ' + sheetNames[i] + '         default   0        := ',file=file2)
        print("[DRC,*,*]         :   ",*Years, sep = "    ",end="",file=file2)
        print("    :=",file=file2) 
#working here !!!!!!!
        Outputsheet_obj=wb_obj[sheetNames[i]]
        Outputsheet_obj2=wb_obj2[sheetNames[i]]

        max_row_Output=Outputsheet_obj.max_row+1
        max_row_Output2=Outputsheet_obj2.max_row+1

        print("max row =",max_row_Output)
        print("max row 2=",max_row_Output2)
        for j in range(2, max_row_Output):
            result = ""
 #           print("j=",j)
 #           print(Outputsheet_obj.cell(row = j, column = 1).value)
            result +=Outputsheet_obj.cell(row = j, column = 1).value 
            for k in range (2,2+Yearnumber):
#                print(j,k)
#                print(Outputsheet_obj.cell(row=34,column=k).value)
                result +=  "   "+ str(Outputsheet_obj.cell(row=j,column=k).value) + "   "    
            print(result,file=file2)
        for j in range(2, max_row_Output2):
            result = ""
 #           print("j=",j)
 #           print(Outputsheet_obj.cell(row = j, column = 1).value)
            result +=Outputsheet_obj2.cell(row = j, column = 1).value 
            for k in range (2,2+Yearnumber):
#                print(j,k)
#                print(Outputsheet_obj.cell(row=34,column=k).value)
                result +=  "   "+ str(Outputsheet_obj2.cell(row=j,column=k).value) + "   "    
            print(result,file=file2)

        print("    ;",file=file2)
        print("#*************",file=file2)
print('here')


#Availibility factor
for i in range(len(sheetNames)):
    if (sheetNames[i] in ['AvailabilityFactor']):
        print('param ' + sheetNames[i] + '         default   1        := ',file=file2)
        print("[DRC,*,*]         :   ",*Years, sep = "    ",end="",file=file2)
        print("    :=",file=file2) 
        Outputsheet_obj=wb_obj[sheetNames[i]]
        Outputsheet_obj2=wb_obj2[sheetNames[i]]
        
        max_row_Output=Outputsheet_obj.max_row
        max_row_Output2=Outputsheet_obj2.max_row

        for i in range(2, max_row_Output+1):
            result = ""
            result +=Outputsheet_obj.cell(row = i, column = 1).value 
            for j in range (2,2+Yearnumber):
                result +=  "   "+ str(Outputsheet_obj.cell(row=i,column=j).value) + "   "    
            print(result,file=file2)
        for i in range(2, max_row_Output2+1):
            result = ""
            result +=Outputsheet_obj2.cell(row = i, column = 1).value 
            for j in range (2,2+Yearnumber):
                result +=  "   "+ str(Outputsheet_obj2.cell(row=i,column=j).value) + "   "    
            print(result,file=file2)

        print("    ;",file=file2)
        print("#*************",file=file2)
print('here')

#Capacity (Variable cost) this has two variable [DRC,*,1,*]
#For mode of opetation 1
Variablesheet_obj=wb_obj['VariableCost']
Variablesheet_obj2=wb_obj2['VariableCost']

print("param     VariableCost	                  default   0	:=     ",file=file2)
print("[DRC,*,1,*]         :   ",*Years, sep = "    ",end='',file=file2)
print("    :=",file=file2) 
#for row in Variablesheet_obj.iter_rows('A'):
#    print("Hello")
max_row_Variablecost = max((c.row for c in Variablesheet_obj['A'] if c.value is not None))
max_row_Variablecost2 = max((c.row for c in Variablesheet_obj2['A'] if c.value is not None))

i=0
for row in Variablesheet_obj.rows:
    i=i+1

    if row[0].value  == 1:

        result = ""
        result +=Variablesheet_obj.cell(row = i, column = 2).value 
        for j in range (3,3+Yearnumber):
            result +=  "   "+ str(Variablesheet_obj.cell(row=i,column=j).value) + "   "    
        print(result,file=file2)
i=0
for row in Variablesheet_obj2.rows:
    i=i+1

    if row[0].value  == 1:

        result = ""
        result +=Variablesheet_obj2.cell(row = i, column = 2).value 
        for j in range (3,3+Yearnumber):
            result +=  "   "+ str(Variablesheet_obj2.cell(row=i,column=j).value) + "   "    
        print(result,file=file2)
print("    ;",file=file2)
print("#*************",file=file2)

#Technology availailability (CapacityFactor)
print("param     CapacityFactor	   default	 1	:=    ",file=file2)
CapacityFactorsheet_obj=wb_obj['CapacityFactor']
CapacityFactorsheet_obj2=wb_obj2['CapacityFactor']

max_row_CapacityFactor = max((c.row for c in CapacityFactorsheet_obj['A'] if c.value is not None))
max_row_CapacityFactor2 = max((c.row for c in CapacityFactorsheet_obj2['A'] if c.value is not None))

for row in  CapacityFactorsheet_obj.iter_rows(min_row=4,max_row=max_row_CapacityFactor ,min_col=5,max_col=5+Yearnumber):
    result = ''
    for cell in row:

        result += " "+ str(cell.value) + " "
    print(result,file=file2)
for row in  CapacityFactorsheet_obj2.iter_rows(min_row=4,max_row=max_row_CapacityFactor2 ,min_col=5,max_col=5+Yearnumber):
    result = ''
    for cell in row:

        result += " "+ str(cell.value) + " "
    print(result,file=file2)


print("    ;",file=file2)
print("#*************",file=file2)


# Emissions penalty
EmissionsPenalty_obj=wb_obj['EmissionsPenalty']
print("param     EmissionsPenalty   	:=     ",file=file2)
print("[DRC,*,*]         :   ",*Years, sep = "    ",end="",file=file2)
print(":=",file=file2)
max_row_EmissionsPenalty = max((c.row for c in EmissionsPenalty_obj['A'] if c.value is not None))
for row in  EmissionsPenalty_obj.iter_rows(min_row=2,max_row=max_row_EmissionsPenalty ,min_col=1,max_col=1+Yearnumber):
    result = ''
    for cell in row:

        result += " "+ str(cell.value) + " "
    print(result,file=file2)
print("    ;",file=file2)
print("#*************",file=file2)


#ReserveMargin, ReserveMarginTagFuel, ReserveMarginTagTechnology
ReserveMargin_obj=wb_obj['ReserveMargin']

print("param     ReserveMargin		:     ",*Years, sep = "    ",end="",file=file2)
print(":=",file=file2)
result=""
for j in range (2,2+Yearnumber):
    result +=  "   "+ str(ReserveMargin_obj.cell(row=2,column=j).value) + "   "    
print("DRC    " +result,file=file2)
print("    ;",file=file2)
print("#*************",file=file2)

for i in range(len(sheetNames)):
    if (sheetNames[i] in ['ReserveMarginTagFuel', 'ReserveMarginTagTechnology']):
        print('param ' + sheetNames[i] + ' default 0 := ',file=file2)
        print("[DRC,*,*]         :   ",*Years, sep = "    ",end='',file=file2)
        print("     := ",file=file2)
        Outputsheet_obj=wb_obj[sheetNames[i]]
        Outputsheet_obj2=wb_obj2[sheetNames[i]]
        max_row_Output=Outputsheet_obj.max_row
        max_row_Output2=Outputsheet_obj2.max_row
        for i in range(2, max_row_Output+1):
            result = ""
            result +=Outputsheet_obj.cell(row = i, column = 1).value 
            for j in range (2,2+Yearnumber):
                result +=  "   "+ str(Outputsheet_obj.cell(row=i,column=j).value) + "   "    
            print(result,file=file2)
        for i in range(2, max_row_Output2+1):
            result = ""
            result +=Outputsheet_obj2.cell(row = i, column = 1).value 
            for j in range (2,2+Yearnumber):
                result +=  "   "+ str(Outputsheet_obj2.cell(row=i,column=j).value) + "   "    
            print(result,file=file2)

        print("    ;",file=file2)
        print("#*************",file=file2)

#OperationalLife, param	CapacityToActivityUnit	default	1	:	:=	;
Techsheet_obj=wb_obj['TECHNOLOGY']
Techsheet_obj2=wb_obj2['TECHNOLOGY']

max_row_Techsheet = max((c.row for c in Techsheet_obj['D'] if c.value is not None))
max_row_Techsheet2 = max((c.row for c in Techsheet_obj2['D'] if c.value is not None))

print(max_row_Techsheet)
for j in range (9,11):
    result = ''
    result += 'param ' + str(Techsheet_obj.cell(row=1,column=j).value)  + ' default  1 :'
    for i in range (2,max_row_Techsheet+1):
        result +=  "   "+ str(Techsheet_obj.cell(row=i,column=4).value) + "   "
    for i in range (2,max_row_Techsheet2+1):
        result +=  "   "+ str(Techsheet_obj2.cell(row=i,column=4).value) + "   "

    result +=  ":=\n"
    result += "DRC"
    for i in range (2,max_row_Techsheet+1):
        result +=  "   "+ str(Techsheet_obj.cell(row=i,column=j).value) + "   "
    for i in range (2,max_row_Techsheet2+1):
        result +=  "   "+ str(Techsheet_obj2.cell(row=i,column=j).value) + "   "    

    print(result,file=file2)
    print("    ;",file=file2)
    print("#*************",file=file2)
#
# Declared upfront for now, but may need to be declared here
#    "param     TotalAnnualMaxCapacity default	99999	:=	;\n",
# TotalAnnualMaxCapacity{r in Region, t in Technology, y in Year} Units: GW
#
#    "param	TotalAnnualMinCapacity default	0	:=	;\n",
# TotalAnnualMinCapacity{r in Region, t in Technology, y in Year} Units: GW
#
#    "param     TotalAnnualMaxCapacityInvestment default	99999	:=	;\n",
# TotalAnnualMaxCapacityInvestment{r in Region, t in Technology, y in Year} Units: GW
#
#    "param	TotalAnnualMinCapacityInvestment default	0	:=	;\n",
# TotalAnnualMinCapacityInvestment{r in Region, t in Technology, y in Year} Units: GW
#
#    "param	TotalTechnologyAnnualActivityUpperLimit	default	99999	:=	;\n",
# param TotalTechnologyAnnualActivityUpperLimit{r in Region, t in Technology, y in Year} Units: PJ
#
#    "param	TotalTechnologyAnnualActivityLowerLimit	default	0	:=	;\n",
# TotalTechnologyAnnualActivityLowerLimit{r in Region, t in Technology, y in Year} Units: PJ
#

#sheetNames = wb_obj.sheetnames already declared don't need to declare again
#print(sheetNames) 
for i in range(len(sheetNames)):
    if (sheetNames[i] in ['AnnualMinCapacity','AnnualMinCapacityInvestment']):
        print('param ' + 'Total'+sheetNames[i] + ' default 0 := ',file=file2)
        print("[DRC,*,*]         :   ",*Years, sep = "    ",end="",file=file2)
        print("     :=",file=file2)
        Outputsheet_obj=wb_obj[sheetNames[i]]
        max_row_Output=Outputsheet_obj.max_row
        Outputsheet_obj2=wb_obj2[sheetNames[i]]
        max_row_Output2=Outputsheet_obj2.max_row

        for i in range(2, max_row_Output+1):
            result = ""
            result +=Outputsheet_obj.cell(row = i, column = 1).value 
            for j in range (2,2+Yearnumber):
                result +=  "   "+ str(Outputsheet_obj.cell(row=i,column=j).value) + "   "
            print(result,file=file2)
        for i in range(2, max_row_Output2+1):
            result = ""
            result +=Outputsheet_obj2.cell(row = i, column = 1).value 
            for j in range (2,2+Yearnumber):
                result +=  "   "+ str(Outputsheet_obj2.cell(row=i,column=j).value) + "   "    

            print(result,file=file2)
        print("    ;",file=file2)
print("#*************",file=file2)
print('here')

for i in range(len(sheetNames)):
    if (sheetNames[i] in ['AnnualMaxCapacity','AnnualMaxCapacityInvestment']):
        print('param ' + 'Total'+sheetNames[i] + ' default 99999 := ',file=file2)
        print("[DRC,*,*]         :   ",*Years, sep = "    ",end="",file=file2)
        print("     :=",file=file2)
        Outputsheet_obj=wb_obj[sheetNames[i]]
        max_row_Output=Outputsheet_obj.max_row
        Outputsheet_obj2=wb_obj2[sheetNames[i]]
        max_row_Output2=Outputsheet_obj2.max_row

        for i in range(2, max_row_Output+1):
            result = ""
            result +=Outputsheet_obj.cell(row = i, column = 1).value 
            for j in range (2,2+Yearnumber):
                result +=  "   "+ str(Outputsheet_obj.cell(row=i,column=j).value) + "   "
            print(result,file=file2)
        for i in range(2, max_row_Output2+1):
            result = ""
            result +=Outputsheet_obj2.cell(row = i, column = 1).value 
            for j in range (2,2+Yearnumber):
                result +=  "   "+ str(Outputsheet_obj2.cell(row=i,column=j).value) + "   "    

            print(result,file=file2)
        print("    ;",file=file2)
print("#*************",file=file2)
print('here')
#param	TotalTechnologyAnnualActivityUpperLimit	default	99999	:= set up in the beginning
#param	TotalTechnologyAnnualActivityLowerLimit	default	0	:= set up in the beginning

# Conversionls{l in TIMESLICE, ls in SEASON}																
# Set equal to 1 to assign a particular time slice to a season. Set equal to 0 in order not to assign a particular time slice to a season.																
print('param    Conversionls              default 	0	 :=	;',file=file2)											
																
# Conversionld{l in TIMESLICE, ld in DAYTYPE}																
# Set equal to 1 to assign a particular time slice to a day type. Set equal to 0 in order not to assign a particular time slice to a day type.																
print('param    Conversionld	          default	0	:=	;',file=file2)

																
# Conversionlh{l in TIMESLICE, lh in DAILYTIMEBRACKET} 																
# Set equal to 1 to assign a particular time slice to a daily time bracket. Set equal to 0 in order not to assign a particular time slice to a daily time bracket.																
print('param    Conversionlh	          default	0	:=	;',file=file2)											
#																
# DaySplit{lh in DAILYTIMEBRACKET, y in YEAR};																
# The length of one time bracket in one specific day as a fraction of the year, e.g., when distinguishing between days and night: 12h/(24h*365d)																
print('param    DaySplit                  default 	0.00137	:=	;',file=file2)											
#																
# TechnologyToStorage{r in REGION, t in TECHNOLOGY, s in STORAGE, m in MODE_OF_OPERATION}																
print('param    TechnologyToStorage	  default	0	:=	;',file=file2)											
#																
# TechnologyFromStorage{r in REGION, t in TECHNOLOGY, s in STORAGE, m in MODE_OF_OPERATION}																
print('param	TechnologyFromStorage	  default	0	:=	;',file=file2)											
#																
# StorageLevelStart{r in REGION, s in STORAGE}																
# At beginning of first year. Attention: if zero, OSeMOSYS will use the first time slices in the entire first day type in the entire first season to fill the storage. 																
# To avoid OSeMOSYS taking a whole part of a season to fill up the storage, and to avoid defining smaller seasons, set it to zero, run the model, and check the StorageLevelYearStart 																
# variable of the following year and use a similar value for StorageLevelStart. Alternatively, model a few years before the first year of your interest.																
print('param 	StorageLevelStart          default 	999	:=	;',file=file2)											
#																
# DaysInDayType{ls in SEASON, ld in DAYTYPE, y in YEAR};																
# Number of days for each day type within a week, i.e., out of 7 																
print('param 	DaysInDayType              default 	7	:=	;',file=file2)											
#																
# StorageMaxChargeRate{r in REGION, s in STORAGE}; Unit: GW																
print('param 	StorageMaxChargeRate        default	99	:=	;',file=file2)											
#																
# StorageMaxDischargeRate{r in REGION, s in STORAGE}; Unit: GW																
print('param 	StorageMaxDischargeRate     default 	99	:=	;',file=file2)											
#																
# MinStorageCharge{r in REGION, s in STORAGE, y in YEAR}; Unit: fraction of MaxStorageCharge, i.e., between 0.00 and 0.99																
print('param 	MinStorageCharge            default	0	:= 	;',file=file2)											
#																
# OperationalLifeStorage{r in REGION, s in STORAGE, y in YEAR}; Unit: years																
print('param 	OperationalLifeStorage      default	99	:=	;',file=file2)											
#																
# CapitalCostStorage{r in REGION, s in STORAGE, y in YEAR}; Unit: USD/GWa																
print('param 	CapitalCostStorage          default 	0	:=	;',file=file2)											
#																
# ResidualStorageCapacity{r in REGION, s in STORAGE, y in YEAR}; 																
# Storage capacity which is available from before the modelling period, or which is know to become available in a specific year. Unit: GWa																
print('param 	ResidualStorageCapacity     default	999	:=	;',file=file2)											
#																
# CapacityOfOneTechnologyUnit{r in REGION, t in TECHNOLOGY, y in YEAR}; Unit: GW																
# Defines the minimum size of one capacity addition. If set to zero, no mixed integer linear programming (MILP) is used and computational time will decrease.																
print('param 	CapacityOfOneTechnologyUnit default	0	:=	;',file=file2)										
#																
# TradeRoute{r in REGION, rr in REGION, f in FUEL, y in YEAR}																
# Defines which region r is linked with which region rr in order to enable or disable trading of a specific fuel. Unit: Fraction, either 1 or 0																
# 1 defines a trade link and 0 ensuring that no trade occurs. Values inbetween are not allowed. If r is linked to rr, rr has also to be linked with r.																
# I.e., for one specific year and fuel, this parameter is entered as a symmetric matrix (with a diagonal of zeros).																
print('param 	TradeRoute                  default	0	:=	;',file=file2)											
#																
print('end;',file=file2)	


#'RETagFuel',max_row_CapacityFactor = max((c.row for c in CapacityFactorsheet_obj['A'] if c.value is not None))
# 'RETagTechnology','ReserveMargin', 'ReserveMarginTagFuel','ReserveMarginTagTechnology', 


file2.close()
print('end')







